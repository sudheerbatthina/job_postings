#!/usr/bin/env python3
"""
AI Engineer Job Finder — v1
===========================
Scrapes job boards (LinkedIn, Indeed, Google, ZipRecruiter) via JobSpy,
scores each posting by keyword relevance + resume match + recency,
and exports a ranked spreadsheet.

v1 scope (per design decision):
  - NO applicant-count filter (skipped; ranks by recency instead)
  - 100% free: no API keys, no proxies required
  - Output: a single .xlsx (+ .csv fallback)

Run:  python ai_engineer_jobs.py
Edit the CONFIG block below to retune.
"""

from __future__ import annotations
import re
import sys
import math
from datetime import datetime, date
from pathlib import Path

import pandas as pd
from jobspy import scrape_jobs

# ============================================================================
# CONFIG  — edit this block, nothing else needed
# ============================================================================

# Titles to search. JobSpy runs one term at a time; we loop and merge.
SEARCH_TERMS = [
    "AI Engineer",
    "Machine Learning Engineer",
    "ML Engineer",
    "Applied Scientist",
    "LLM Engineer",
    "Generative AI Engineer",
    "Deep Learning Engineer",
]

LOCATION       = "United States"   # any city/state/country string
IS_REMOTE      = False             # True = remote only; False = all
COUNTRY_INDEED = "USA"             # required by Indeed/Glassdoor
SITES          = ["linkedin", "indeed", "google", "zip_recruiter"]
RESULTS_WANTED = 40                # per site, per search term
HOURS_OLD      = 168               # recency window (168h = 7 days)

# Optional resume for personalized scoring. .txt or .pdf. Leave "" to skip.
RESUME_PATH    = ""                # e.g. "resume.pdf"

# How much each signal counts toward the final rank (need not sum to 1).
WEIGHTS = {
    "keyword": 0.45,   # how AI-engineer-relevant the posting is
    "resume":  0.25,   # overlap with your resume (0 if no resume given)
    "recency": 0.30,   # newer = higher
}

# A job must clear this keyword score (0-1) to be kept at all — kills noise.
MIN_KEYWORD_SCORE = 0.10

# Final sort: "composite" (default), "recency", or "keyword".
SORT_MODE = "composite"

OUTPUT_DIR = Path(".")
# ============================================================================

# Weighted relevance vocabulary. Higher weight = stronger AI-eng signal.
AI_KEYWORDS = {
    # core role signals
    "machine learning": 3, "deep learning": 3, "ai engineer": 4,
    "ml engineer": 4, "applied scientist": 3, "llm": 3, "large language model": 3,
    "generative ai": 3, "genai": 3, "mlops": 3, "neural network": 2,
    # frameworks / tooling
    "pytorch": 2, "tensorflow": 2, "hugging face": 2, "huggingface": 2,
    "langchain": 2, "transformers": 2, "scikit-learn": 1, "keras": 1,
    "vector database": 2, "rag": 2, "retrieval augmented": 2, "fine-tuning": 2,
    "fine tuning": 2, "embeddings": 2, "diffusion": 2, "computer vision": 2,
    "nlp": 2, "natural language processing": 2, "reinforcement learning": 2,
    # platform / infra
    "kubernetes": 1, "docker": 1, "aws": 1, "gcp": 1, "azure": 1,
    "spark": 1, "airflow": 1, "model serving": 2, "inference": 2,
    "feature store": 1, "data pipeline": 1, "python": 1,
}
# Maximum reachable raw weight (used to normalize keyword score to 0-1).
_KW_NORM = sum(sorted(AI_KEYWORDS.values(), reverse=True)[:8])  # top-8 cap

# Titles containing these are almost never an IC AI-eng role -> drop.
TITLE_BLOCKLIST = [
    "recruit", "sales", "account executive", "intern,", "internship",
    "professor", "lecturer", "teacher", "marketing", "customer success",
]

STOPWORDS = set("""a an the and or of to in for with on at by from as is are be this that
your you we our will have has can able role team work years experience strong
skills ability requirements preferred plus etc using use used job position company""".split())


def load_resume(path: str) -> set[str]:
    """Return a set of meaningful resume tokens, or empty set if no resume."""
    if not path:
        return set()
    p = Path(path)
    if not p.exists():
        print(f"  ! resume not found at {path} — scoring without it")
        return set()
    text = ""
    if p.suffix.lower() == ".pdf":
        try:
            import pdfplumber
            with pdfplumber.open(p) as pdf:
                text = "\n".join(pg.extract_text() or "" for pg in pdf.pages)
        except Exception as e:
            print(f"  ! could not read PDF resume ({e}) — install pdfplumber")
            return set()
    else:
        text = p.read_text(errors="ignore")
    toks = re.findall(r"[a-zA-Z][a-zA-Z+#.\-]{2,}", text.lower())
    return {t for t in toks if t not in STOPWORDS}


def keyword_score(title: str, desc: str) -> float:
    """0-1 relevance from weighted keyword hits (title hits weighted 3x)."""
    t, d = (title or "").lower(), (desc or "").lower()
    raw = 0
    for kw, w in AI_KEYWORDS.items():
        if kw in t:
            raw += w * 3
        elif kw in d:
            raw += w
    return min(1.0, raw / (_KW_NORM * 3))


def resume_score(desc: str, resume_tokens: set[str]) -> float:
    """0-1 overlap: share of distinct job tokens also present in the resume."""
    if not resume_tokens or not desc:
        return 0.0
    job_toks = {t for t in re.findall(r"[a-zA-Z][a-zA-Z+#.\-]{2,}", desc.lower())
                if t not in STOPWORDS}
    if not job_toks:
        return 0.0
    return len(job_toks & resume_tokens) / len(job_toks)


def recency_score(d, window_hours: int) -> float:
    """1.0 = posted today, decaying to ~0 at the edge of the window."""
    if d is None or (isinstance(d, float) and math.isnan(d)):
        return 0.3  # unknown date -> mild penalty, not zero
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d).date()
        except ValueError:
            return 0.3
    if isinstance(d, datetime):
        d = d.date()
    days_old = (date.today() - d).days
    window_days = max(1, window_hours / 24)
    return max(0.0, 1.0 - days_old / window_days)


def title_blocked(title: str) -> bool:
    t = (title or "").lower()
    return any(b in t for b in TITLE_BLOCKLIST)


def scrape_all() -> pd.DataFrame:
    frames = []
    for term in SEARCH_TERMS:
        print(f"  · scraping: {term!r}")
        try:
            df = scrape_jobs(
                site_name=SITES,
                search_term=term,
                google_search_term=f"{term} jobs in {LOCATION}",
                location=LOCATION,
                results_wanted=RESULTS_WANTED,
                hours_old=HOURS_OLD,
                country_indeed=COUNTRY_INDEED,
                is_remote=IS_REMOTE,
                linkedin_fetch_description=True,  # needed for resume/keyword scoring
                description_format="markdown",
                verbose=0,
            )
            if df is not None and len(df):
                df["search_term"] = term
                frames.append(df)
        except Exception as e:
            print(f"    ! {term} failed: {e}")
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def score_and_rank(df: pd.DataFrame, resume_tokens: set[str]) -> pd.DataFrame:
    if df.empty:
        return df
    # dedupe on the job URL (fall back to title+company)
    df = df.drop_duplicates(subset="job_url")
    df = df.drop_duplicates(subset=["title", "company"]).reset_index(drop=True)

    df = df[~df["title"].apply(title_blocked)].copy()

    df["kw_score"]      = df.apply(lambda r: keyword_score(r.get("title"), r.get("description")), axis=1)
    df["resume_match"]  = df.apply(lambda r: resume_score(r.get("description"), resume_tokens), axis=1)
    df["recency"]       = df["date_posted"].apply(lambda d: recency_score(d, HOURS_OLD))

    df = df[df["kw_score"] >= MIN_KEYWORD_SCORE].copy()

    df["score"] = (
        WEIGHTS["keyword"] * df["kw_score"]
        + WEIGHTS["resume"] * df["resume_match"]
        + WEIGHTS["recency"] * df["recency"]
    ).round(3)
    # interpretable 0-100: normalize by max possible weighted sum
    total_w = sum(WEIGHTS.values())
    df["score_100"] = (df["score"] / total_w * 100).round(0).astype(int)

    sort_col = {"recency": "recency", "keyword": "kw_score"}.get(SORT_MODE, "score")
    df = df.sort_values([sort_col, "recency"], ascending=False).reset_index(drop=True)
    df.insert(0, "rank", range(1, len(df) + 1))
    return df


def export(df: pd.DataFrame) -> Path:
    stamp = datetime.now().strftime("%Y%m%d")
    cols = ["rank", "score_100", "title", "company", "location",
            "date_posted", "recency", "kw_score", "resume_match",
            "is_remote", "min_amount", "max_amount", "job_url", "search_term"]
    out = df[[c for c in cols if c in df.columns]].copy()
    out = out.rename(columns={
        "score_100": "score", "min_amount": "salary_min",
        "max_amount": "salary_max", "job_url": "apply_url",
        "kw_score": "keyword_fit", "resume_match": "resume_fit",
    })

    xlsx_path = OUTPUT_DIR / f"ai_engineer_jobs_{stamp}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xl:
        out.to_excel(xl, index=False, sheet_name="AI Engineer Jobs")
        ws = xl.sheets["AI Engineer Jobs"]
        # freeze header, auto-width, make apply_url clickable
        ws.freeze_panes = "A2"
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(out.columns, 1):
            width = max(len(str(col)), out[col].astype(str).str.len().max() if len(out) else 10)
            ws.column_dimensions[get_column_letter(i)].width = min(60, width + 2)
        url_col = list(out.columns).index("apply_url") + 1
        for row in range(2, len(out) + 2):
            cell = ws.cell(row=row, column=url_col)
            if cell.value:
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    out.to_csv(OUTPUT_DIR / f"ai_engineer_jobs_{stamp}.csv", index=False)
    return xlsx_path


def main():
    print("AI Engineer Job Finder — v1\n")
    resume_tokens = load_resume(RESUME_PATH)
    if resume_tokens:
        print(f"  resume loaded: {len(resume_tokens)} distinct tokens\n")
    raw = scrape_all()
    print(f"\n  scraped {len(raw)} raw rows")
    ranked = score_and_rank(raw, resume_tokens)
    if ranked.empty:
        print("  no jobs passed filtering. Loosen MIN_KEYWORD_SCORE or widen HOURS_OLD.")
        sys.exit(0)
    path = export(ranked)
    print(f"  kept {len(ranked)} ranked jobs")
    print(f"  -> {path}")
    print("\n  top 5:")
    for _, r in ranked.head().iterrows():
        print(f"    [{r['score_100']:>3}] {r['title'][:45]:45} {str(r['company'])[:25]}")


if __name__ == "__main__":
    main()
