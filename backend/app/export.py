"""Builds a ranked-results spreadsheet in memory (no disk writes)."""

from __future__ import annotations
import io

import pandas as pd
from openpyxl.utils import get_column_letter


COLUMN_ORDER = ["rank", "ats_score", "title", "company", "location",
                "date_posted", "is_remote", "min_amount", "max_amount",
                "missing_keywords", "job_url", "search_term"]

RENAME = {
    "ats_score": "score", "min_amount": "salary_min",
    "max_amount": "salary_max", "job_url": "apply_url",
}


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    out = df[[c for c in COLUMN_ORDER if c in df.columns]].copy()
    out = out.rename(columns=RENAME)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xl:
        out.to_excel(xl, index=False, sheet_name="AI Engineer Jobs")
        ws = xl.sheets["AI Engineer Jobs"]
        ws.freeze_panes = "A2"
        for i, col in enumerate(out.columns, 1):
            width = max(len(str(col)), out[col].astype(str).str.len().max() if len(out) else 10)
            ws.column_dimensions[get_column_letter(i)].width = min(60, width + 2)
        if "apply_url" in out.columns:
            url_col = list(out.columns).index("apply_url") + 1
            for row in range(2, len(out) + 2):
                cell = ws.cell(row=row, column=url_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
    return buf.getvalue()
